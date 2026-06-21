#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""亚马逊交易汇总报表 PDF 批量提取工具 v5"""
import argparse
import os, re
import sys
from pathlib import Path
try:
    import pdfplumber
except ImportError:
    pdfplumber = None
from pypdf import PdfReader
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PROJECT_DIR = Path(__file__).resolve().parents[1]
MONTHLY_REPORT_DIR = PROJECT_DIR / "亚马逊汇总报表-PDF（按月）"
BASE_DIR = str(MONTHLY_REPORT_DIR if MONTHLY_REPORT_DIR.is_dir() else PROJECT_DIR)
OUTPUT_PATH = os.path.join(BASE_DIR, "outputs", "亚马逊交易汇总报表_汇总.xlsx")

COUNTRY_NAME_TO_CODE = {
    "美国": "US", "加拿大": "CA", "英国": "UK", "德国": "DE", "法国": "FR",
    "意大利": "IT", "西班牙": "ES", "荷兰": "NL", "比利时": "BE",
    "爱尔兰": "IE", "波兰": "PL", "瑞典": "SE", "墨西哥": "MX",
    "沙特": "SA", "阿联酋": "AE", "日本": "JP", "澳洲": "AU", "澳大利亚": "AU",
}
CODE_TO_COUNTRY_NAME = {}
for _country_name, _country_code in COUNTRY_NAME_TO_CODE.items():
    CODE_TO_COUNTRY_NAME.setdefault(_country_code, _country_name)

FIELD_CN = {
    "Income":"收入合计","Expenses":"费用合计","Tax":"税款合计","Transfers":"转账合计",
    "Seller fulfilled product sales":"卖家自发货商品销售额",
    "Product sales (non-FBA)":"卖家自发货商品销售额",
    "Seller fulfilled product sale refunds":"卖家自发货商品退款",
    "Product sale refunds (non-FBA)":"卖家自发货商品退款",
    "Product sale refunds(non-FBA)":"卖家自发货商品退款",
    "FBA product sales":"FBA商品销售额",
    "FBA product sale refunds":"FBA商品退款",
    "FBA inventory credit":"FBA库存信用",
    "FBA liquidation proceeds":"FBA清仓收益",
    "FBA Liquidations proceeds adjustments":"FBA清仓收益调整",
    "Shipping credits":"运费补贴收入",
    "Shipping credit refunds":"运费补贴退款",
    "Gift wrap credits":"礼品包装费收入",
    "Gift wrap credit refunds":"礼品包装费退款",
    "Promotional rebates":"促销折扣",
    "Promotional rebate refunds":"促销折扣退款",
    "A-to-z Guarantee claims":"A-to-Z保障索赔",
    "Chargebacks":"信用卡拒付",
    "Amazon Shipping Reimbursement Adjustments":"亚马逊配送报销调整",
    "SAFE-T reimbursement":"SAFE-T赔偿",
    "Receivables Deductions":"应收账款扣减",
    "Receivables Reversals":"应收账款冲回",
    "Amazon Shipping Charge Adjustments":"亚马逊配送费调整",
    "Penalty Charges":"罚款费用",
    "Commingling VAT":"混合库存增值税",
    "Seller fulfilled selling fees":"卖家自发货销售佣金",
    "FBA selling fees":"FBA销售佣金",
    "Selling fee refunds":"销售佣金退款",
    "FBA transaction fees":"FBA交易费",
    "FBA transaction fee refunds":"FBA交易费退款",
    "Other transaction fees":"其他交易费",
    "Other transaction fee refunds":"其他交易费退款",
    "FBA inventory and inbound services fees":"FBA库存及入库服务费",
    "Shipping label purchases":"运单购买费",
    "Shipping label refunds":"运单退款",
    "Carrier shipping label adjustments":"承运商运单调整",
    "Service fees":"服务费",
    "Refund administration fees":"退款处理费",
    "Adjustments":"其他调整",
    "Cost of Advertising":"广告费用",
    "Refund for Advertiser":"广告退款",
    "Liquidations fees":"清仓手续费",
    "Transfers to bank account":"转账至银行账户",
    "Failed transfers to bank account":"银行转账失败",
    "Charges to credit card and other debt recovery":"信用卡扣款及其他债务回收",
    "Disburse to Amazon Gift Card balance":"转入亚马逊礼品卡余额",
    "Product, shipping and gift wrap taxes collected":"商品/运费/礼品包装税（已收）",
    "Product, shipping and gift wrap taxes refunded":"商品/运费/礼品包装税（已退）",
    "Amazon obligated tax withheld":"亚马逊代扣税款",
    "Amazon Obligated Tax and Regulatory Fee Withheld":"亚马逊代扣税款及监管费",
    "Product, shipping, gift wrap taxes and regulatory fee collected":"商品/运费/礼品包装税及监管费（已收）",
    "Product, shipping, gift wrap taxes and regulatory fee refunded":"商品/运费/礼品包装税及监管费（已退）",
    "Product VAT adjustment":"商品增值税调整",
    "Income tax withheld":"预扣所得税",
    "Others":"其他",
    "Amazon shipping fee refunds":"亚马逊配送服务费退款",
    "Cost of points granted":"积分授予成本",
    "Cost of points returned":"积分返还成本",
    "Low Value Goods Tax":"低价值商品税",
}

DE_TO_EN = {
    "Einnahmen":"Income","Ausgaben":"Expenses","Steuer":"Tax","Übertragungen":"Transfers",
    "Verkäufe, die durch Verkäufer selbst verschickt wurden":"Seller fulfilled product sales",
    "Erstattungen für vom Verkäufer versandte Artikel":"Seller fulfilled product sale refunds",
    "Verkäufe mit Versand durch Amazon":"FBA product sales",
    "Erstattungen für durch Amazon versandte Artikel":"FBA product sale refunds",
    "FBA Lagerbestandguthaben":"FBA inventory credit",
    "Erlöse der Liquidation über Versand durch Amazon":"FBA liquidation proceeds",
    "Anpassungen des \"Versand durch Amazon\" -Liquidationserlöses":"FBA Liquidations proceeds adjustments",
    "Versandkostengutschriften":"Shipping credits",
    "Erstattungen zu Versandkostengutschriften":"Shipping credit refunds",
    "Gutschrift für Geschenkverpackung":"Gift wrap credits",
    "Erstattungen zu Gutschriften aus Geschenkverpackungen":"Gift wrap credit refunds",
    "Werbeaktions-Rabatte":"Promotional rebates",
    "Erstattungen zu Werbeaktions-Rabatt":"Promotional rebate refunds",
    "A-bis-Z-Garantieanträge":"A-to-z Guarantee claims",
    "Rückbuchungen":"Chargebacks",
    "Amazon Shipping Reimbursement Adjustments":"Amazon Shipping Reimbursement Adjustments",
    "SAFE-T-Erstattung":"SAFE-T reimbursement",
    "Commingling VAT":"Commingling VAT",
    "Verkaufsgebühren Versand durch Verkäufer":"Seller fulfilled selling fees",
    "Verkaufsgebühren Versand durch Amazon":"FBA selling fees",
    "Erstattungen zur Verkaufsgebühr":"Selling fee refunds",
    "Transaktionsgebühren Versand durch Amazon":"FBA transaction fees",
    "Erstattungen zur Transaktionsgebühr - Versand durch Amazon":"FBA transaction fee refunds",
    "Andere Transaktionsgebühren":"Other transaction fees",
    "Sonstige Erstattungen zu Transaktionsgebühren":"Other transaction fee refunds",
    "Lagerbestands- und Service-Gebühren für Versand durch Amazon":"FBA inventory and inbound services fees",
    "Erwerb von Versandscheinen":"Shipping label purchases",
    "Erstattungen zu Versandscheinen":"Shipping label refunds",
    "Anpassungen des Versandscheins":"Carrier shipping label adjustments",
    "Servicegebühren":"Service fees",
    "Bearbeitungsgebühren für Erstattungen":"Refund administration fees",
    "Anpassungen":"Adjustments",
    "Werbekosten":"Cost of Advertising",
    "Gutschrift für Inserenten":"Refund for Advertiser",
    "Gebühren für Liquidationen":"Liquidations fees",
    "Überweisungen auf Bankkonto":"Transfers to bank account",
    "Fehlgeschlagene Überweisungen zu Bankkonto":"Failed transfers to bank account",
    "Kreditkartengebühren und andere Inkassogebühren":"Charges to credit card and other debt recovery",
    "Eingezogene Produkt-, Versand- und Geschenkverpackungssteuern":"Product, shipping and gift wrap taxes collected",
    "Zurückerstattete Produkt-, Versand- und Geschenkverpackungssteuern":"Product, shipping and gift wrap taxes refunded",
    "Amazon verpflichtete einbehaltene Steuer":"Amazon obligated tax withheld",
    "Zwischensummen":"subtotals",
}

SECTION_EN = {"income","expenses","tax","transfers"}
SECTION_DE = {"einnahmen","ausgaben","steuer","übertragungen"}
SECTION_SE = {"intäkter","utgifter","moms","överföringar"}
SECTION_NL = {"inkomen","uitgaven","belasting","overboekingen"}
SECTION_JP = {"収入","支出","税金","振込み"}
SECTION_MX = {"ingresos","gastos","impuesto","transferencias"}
SECTION_FR = {"revenus","dépenses","taxe","transferts"}
SECTION_IT = {"ricavi","spese","imposte","trasferimenti"}
SECTION_PL = {"przychód","wydatki","podatek","przelewy"}
SECTION_ES = {"ingresos","costes","impuesto","transferencias"}
ALL_SECTIONS = SECTION_EN | SECTION_DE | SECTION_SE | SECTION_NL | SECTION_JP | SECTION_MX | SECTION_FR | SECTION_IT | SECTION_PL | SECTION_ES
SUBTOTAL_WORDS = {"subtotals","subtotal","zwischensummen","zwischensumme","delsummor","delsumma","小計","subtotales","subtotale","totaux partiels","subtotalen","sous-totaux","totali parziali","小计","sumy czciowe","totale parziale"}
HEADER_WORDS = {"debits","credits","belastungen","einnahmen","debiteringar","krediter","debetbetalingen","tegoeden","débitos","créditos","支払額","入金額","débits","crédits","addebiti","accrediti","obcienia","noty kredytowe","débitos abonos","abonos",
                "débits crédits","addebiti accrediti","debetbetalingen tegoeden","debiteringar krediter","débitos créditos","obcienia noty kredytowe","支払額 入金額"}
SKIP_STARTS = ["page ","sida ","pagina ","página ","ページ","questions?","information in","diese informationen",
               "alle gebühren","can include","kann transaktionen","account types",
               "account activity","vorgänge vom","kontoaktivitet","accountactiviteit",
               "aktywno","actividad de","display name","legal name","visningsnamn",
               "juridiskt","weergavenaam","wettelijke","wywietlana","nombre para",
               "nombre legal","表示名","正式名称","anzeigename","eingetragener",
               "all amounts","alle beträge","alla belopp","alle bedragen",
               "summaries","zusammenfassungen","sammanfattningar","overzichten",
               "podsumowania","resúmenes","概要","totals","gesamt","totalsummor",
               "totalen","totales","合計","=====","-----",
               "informationen i","de informatie","informacja","la información",
               "本ページ","les informations","le informazioni","informacje zawarte",
               "tous les frais","tutti i costi","synthèses","sintesi","podsumowania",
               "résumés","nom affiché","nom commercial","nom légal","dénomination",
               "nome visualizzato","nome legale","wywietlana","nazwa prawna",
               "activité du compte","attività dell","aktywno","actividad de la cuenta","pág.","nombre público","razón social","resúmenes","todas las tarifas",
               "all fees listed","informatie op dit afschrift"]

# ── 瑞典语 → 英语 ──
SE_TO_EN = {
    "Intäkter":"Income","Utgifter":"Expenses","Moms":"Tax","Överföringar":"Transfers",
    "Produktförsäljning (ej FBA)":"Seller fulfilled product sales",
    "Återbetalningar av produktförsäljning (ej FBA)":"Seller fulfilled product sale refunds",
    "FBA-produktförsäljning":"FBA product sales",
    "Återbetalningar FBA-produktförsäljning":"FBA product sale refunds",
    "Kredit för FBA-lager":"FBA inventory credit",
    "Intäkter från FBA-utförsäljning":"FBA liquidation proceeds",
    "FBA Liquidations proceeds adjustments":"FBA Liquidations proceeds adjustments",
    "Fraktkrediter":"Shipping credits",
    "Återbetalningar av fraktkredit":"Shipping credit refunds",
    "Krediter presentinslagning":"Gift wrap credits",
    "Kreditåterbetalningar presentinslagning":"Gift wrap credit refunds",
    "Kampanjrabatter":"Promotional rebates",
    "Återbetalning av kampanjrabatt":"Promotional rebate refunds",
    "Anspråk på A-to-z guarantee":"A-to-z Guarantee claims",
    "Återdebiteringar":"Chargebacks",
    "Återbetalning av Amazon-frakt":"Amazon Shipping Reimbursement Adjustments",
    "SAFE-T-återbetalning":"SAFE-T reimbursement",
    "Varierande moms":"Commingling VAT",
    "Återkallande av fordringar":"Receivables Reversals",
    "Amazon fraktkostnader":"Amazon Shipping Charge Adjustments",
    "Avdrag för fordringar":"Receivables Deductions",
    "Försäljningsavgifter levererat av säljare":"Seller fulfilled selling fees",
    "FBA-försäljningsavgifter":"FBA selling fees",
    "Återbetalningar av försäljningsavgift":"Selling fee refunds",
    "FBA-transaktionsavgifter":"FBA transaction fees",
    "Återbetalningar av FBA-transaktionsavgift":"FBA transaction fee refunds",
    "Övriga transaktionsavgifter":"Other transaction fees",
    "Återbetalningar av andra transaktionsavgifter":"Other transaction fee refunds",
    "FBA-avgifter för lagerhållning och inkommande tjänster":"FBA inventory and inbound services fees",
    "Inköp fraktetikett":"Shipping label purchases",
    "Återbetalningar fraktetikett":"Shipping label refunds",
    "Justeringar av fraktetikett speditör":"Carrier shipping label adjustments",
    "Serviceavgifter":"Service fees",
    "Administrativ avgift för återbetalning":"Refund administration fees",
    "Justeringar":"Adjustments",
    "Reklamkostnad":"Cost of Advertising",
    "Återbetalning för annonsör":"Refund for Advertiser",
    "Avgifter för utförsäljningar":"Liquidations fees",
    "Överföringar till bankkonto":"Transfers to bank account",
    "Misslyckade överföringar till bankkonto":"Failed transfers to bank account",
    "Avgifter för kreditkort och annan skuldindrivning":"Charges to credit card and other debt recovery",
    "Insamlade skatter för produkter, frakt och presentinslagning":"Product, shipping and gift wrap taxes collected",
    "Återbetalade skatter för produkter, frakt och presentinslagning":"Product, shipping and gift wrap taxes refunded",
    "Uttagen obligatorisk Amazon-skatt":"Amazon obligated tax withheld",
    "delsummor":"subtotals","delsumma":"subtotals",
}

# ── 荷兰语 → 英语 ──
NL_TO_EN = {
    "Inkomen":"Income","Uitgaven":"Expenses","Belasting":"Tax","Overboekingen":"Transfers",
    "Verkoop van producten (niet-FBA)":"Seller fulfilled product sales",
    "Terugbetalingen voor productverkoop (niet-FBA)":"Seller fulfilled product sale refunds",
    "Verkoop van FBA-producten":"FBA product sales",
    "Terugbetalingen voor FBA-productverkoop":"FBA product sale refunds",
    "FBA-voorraadkrediet":"FBA inventory credit",
    "FBA-liquidatieopbrengsten":"FBA liquidation proceeds",
    "Aanpassingen FBA-liquidatieopbrengsten":"FBA Liquidations proceeds adjustments",
    "Verzendkrediet":"Shipping credits",
    "Terugbetalingen voor verzendkrediet":"Shipping credit refunds",
    "Cadeauverpakkingskrediet":"Gift wrap credits",
    "Terugbetalingen voor cadeauverpakkingskrediet":"Gift wrap credit refunds",
    "Promotionele kortingen":"Promotional rebates",
    "Terugbetalingen voor promotionele kortingen":"Promotional rebate refunds",
    "A-to-z-garantieclaims":"A-to-z Guarantee claims",
    "Terugboekingen":"Chargebacks",
    "Aanpassingen Amazon-verzendvergoeding":"Amazon Shipping Reimbursement Adjustments",
    "SAFE-T-vergoeding":"SAFE-T reimbursement",
    "Commingling btw":"Commingling VAT",
    "Verkoper heeft verkoopkosten voldaan":"Seller fulfilled selling fees",
    "FBA-verkoopkosten":"FBA selling fees",
    "Terugbetalingen voor verkoopkosten":"Selling fee refunds",
    "FBA-transactiekosten":"FBA transaction fees",
    "Terugbetalingen voor FBA-transactiekosten":"FBA transaction fee refunds",
    "Overige transactiekosten":"Other transaction fees",
    "Terugbetalingen voor overige transactiekosten":"Other transaction fee refunds",
    "FBA-kosten voor voorraad en inkomende services":"FBA inventory and inbound services fees",
    "Aankopen van verzendetiketten":"Shipping label purchases",
    "Terugbetalingen voor verzendetiketten":"Shipping label refunds",
    "Aanpassingen verzendetiket vervoerder":"Carrier shipping label adjustments",
    "Servicekosten":"Service fees",
    "Administratiekosten voor terugbetalingen":"Refund administration fees",
    "Aanpassingen":"Adjustments",
    "Advertentiekosten":"Cost of Advertising",
    "Terugbetaling voor adverteerder":"Refund for Advertiser",
    "Liquidatiekosten":"Liquidations fees",
    "Overboekingen naar bankrekening":"Transfers to bank account",
    "Mislukte overboekingen naar bankrekening":"Failed transfers to bank account",
    "Kosten voor creditcard en andere schuldinvordering":"Charges to credit card and other debt recovery",
    "Geïnde product-, verzend- en cadeauverpakkingsbelastingen":"Product, shipping and gift wrap taxes collected",
    "Terugbetaalde product-, verzend- en cadeauverpakkingsbelastingen":"Product, shipping and gift wrap taxes refunded",
    "Ingehouden verplichte Amazon-belasting":"Amazon obligated tax withheld",
    "Terugbetalingen voor verkoop van FBA-producten":"FBA product sale refunds",
    "Verzendtegoeden":"Shipping credits",
    "Terugbetalingen verzendtegoed":"Shipping credit refunds",
    "Tegoeden cadeauverpakking":"Gift wrap credits",
    "Restituties cadeauverpakking":"Gift wrap credit refunds",
    "Promotiekortingen":"Promotional rebates",
    "Terugbetalingen promotiekorting":"Promotional rebate refunds",
    "A-tot-Z-garantieclaims":"A-to-z Guarantee claims",
    "Terugvorderingen":"Chargebacks",
    "Terugbetaling van Amazon-verzendkosten":"Amazon Shipping Reimbursement Adjustments",
    "SAFE-T-terugbetaling":"SAFE-T reimbursement",
    "Terugboekingen van vorderingen":"Receivables Reversals",
    "Verzendkosten van Amazon":"Amazon Shipping Charge Adjustments",
    "Inhoudingen op vorderingen":"Receivables Deductions",
    "Terugbetaling van verkoopkosten":"Selling fee refunds",
    "Terugbetaling FBA-transactiekosten":"FBA transaction fee refunds",
    "Restituties overige transactiekosten":"Other transaction fee refunds",
    "Kosten voor voorraad en binnenkomende goederen bij FBA":"FBA inventory and inbound services fees",
    "Aankopen verzendetiketten":"Shipping label purchases",
    "Terugbetalingen verzendetiket":"Shipping label refunds",
    "Aanpassingen aan verzendetiketten van de transportdienst":"Carrier shipping label adjustments",
    "Administratiekosten terugbetalen":"Refund administration fees",
    "Mislukte overschrijvingen naar bankrekening":"Failed transfers to bank account",
    "Kosten voor creditcardbetalingen en andere incasso\'s":"Charges to credit card and other debt recovery",
    "Product-, verzend- en cadeauverpakkingsbelasting verzameld":"Product, shipping and gift wrap taxes collected",
    "Product-, verzend- en cadeauverpakkingsbelasting terugbetaald":"Product, shipping and gift wrap taxes refunded",
    "Verplichte belasting ingehouden door Amazon":"Amazon obligated tax withheld",
    "subtotalen":"subtotals",
}

# ── 日语 → 英语 ──
JP_TO_EN = {
    "収入":"Income","支出":"Expenses","税金":"Tax","振込み":"Transfers",
    "自社出荷商品の売上額":"Seller fulfilled product sales",
    "自社出荷による商品の返金額":"Seller fulfilled product sale refunds",
    "FBA ご利用商品の売上額":"FBA product sales",
    "FBA ご利用商品の返金":"FBA product sale refunds",
    "FBA在庫クレジット":"FBA inventory credit",
    "FBA在庫クリアランスプログラムの利益":"FBA liquidation proceeds",
    "配送料":"Shipping credits",
    "配送料の返金額":"Shipping credit refunds",
    "ギフト包装料":"Gift wrap credits",
    "ギフト包装料の返金額":"Gift wrap credit refunds",
    "プロモーション割引額":"Promotional rebates",
    "プロモーション割引の返金額":"Promotional rebate refunds",
    "Amazonマーケットプレイス保証申請":"A-to-z Guarantee claims",
    "チャージバック":"Chargebacks",
    "SAFE-T補てん":"SAFE-T reimbursement",
    "マーケットプレイス配送サービス料金の調整":"Amazon Shipping Reimbursement Adjustments",
    "自社出荷商品の出品手数料":"Seller fulfilled selling fees",
    "FBA ご利用商品の出品手数料":"FBA selling fees",
    "出品手数料の返金額":"Selling fee refunds",
    "FBA 配送代行手数料":"FBA transaction fees",
    "FBA 配送代行手数料の返金額":"FBA transaction fee refunds",
    "取引に関するその他の手数料":"Other transaction fees",
    "取引に関するその他の手数料の返金額":"Other transaction fee refunds",
    "FBA 在庫保管と納品に関する手数料":"FBA inventory and inbound services fees",
    "月額登録料など":"Service fees",
    "返金処理手数料":"Refund administration fees",
    "Amazon から請求または、振り込まれる、その他の金額":"Adjustments",
    "広告費用":"Cost of Advertising",
    "広告費の返金":"Refund for Advertiser",
    "在庫清算の手数料":"Liquidations fees",
    "マーケットプレイス配送サービス料金":"Amazon Shipping Charge Adjustments",
    "マーケットプレイス配送サービス料金の返金":"Amazon shipping fee refunds",
    "付与したポイントの費用":"Cost of points granted",
    "返却されたポイントの費用":"Cost of points returned",
    "銀行口座への振込額":"Transfers to bank account",
    "銀行口座への振込が完了していない金額":"Failed transfers to bank account",
    "クレジットカードおよびその他の債務回収への請求":"Charges to credit card and other debt recovery",
    "商品、配送、ギフト包装に対して税金が徴収されました":"Product, shipping and gift wrap taxes collected",
    "商品、配送、ギフト包装に対して税金還付されました":"Product, shipping and gift wrap taxes refunded",
    "Amazonの源泉徴収":"Amazon obligated tax withheld",
    "小計":"subtotals",
}

# ── 西班牙语(MX) → 英语 ──
MX_TO_EN = {
    "Ingresos":"Income","Gastos":"Expenses","Impuesto":"Tax","Transferencias":"Transfers",
    "Venta de producto realizada por el vendedor":"Seller fulfilled product sales",
    "Reembolsos de venta de producto realizada por el vendedor":"Seller fulfilled product sale refunds",
    "Ventas de producto FBA":"FBA product sales",
    "Reembolsos de venta de producto FBA":"FBA product sale refunds",
    "Crédito de inventario del FBA":"FBA inventory credit",
    "Créditos de envío":"Shipping credits",
    "Reembolsos de crédito de envío":"Shipping credit refunds",
    "Créditos por envoltorio de regalo":"Gift wrap credits",
    "Reembolsos de crédito de envoltorio de regalo":"Gift wrap credit refunds",
    "Descuentos promocionales":"Promotional rebates",
    "Reembolsos de descuento promocional":"Promotional rebate refunds",
    "Reclamos de A-to-z Guarantee":"A-to-z Guarantee claims",
    "Reintegros":"Chargebacks",
    "Reembolso por envío de Amazon":"Amazon Shipping Reimbursement Adjustments",
    "SAFE-T reimbursement":"SAFE-T reimbursement",
    "Tarifas de venta realizadas por el vendedor":"Seller fulfilled selling fees",
    "Tarifas de venta FBA":"FBA selling fees",
    "Reembolsos de tarifa de venta":"Selling fee refunds",
    "Tarifas de transacción FBA":"FBA transaction fees",
    "Reembolsos de tarifas de transacción FBA":"FBA transaction fee refunds",
    "Tarifas de otra transacción":"Other transaction fees",
    "Reembolsos de tarifas de otras transacciones":"Other transaction fee refunds",
    "Tarifas de inventario y de servicios de Logística de Amazon":"FBA inventory and inbound services fees",
    "Compras de etiqueta de envío":"Shipping label purchases",
    "Reembolsos de etiqueta de envío":"Shipping label refunds",
    "Ajustes en la etiqueta de envío del transportista":"Carrier shipping label adjustments",
    "Tarifas de servicio":"Service fees",
    "Tarifas de administración de reembolso":"Refund administration fees",
    "Ajustes":"Adjustments",
    "Costo de la publicidad":"Cost of Advertising",
    "Reembolso para el promotor":"Refund for Advertiser",
    "Transferencias a cuenta bancaria":"Transfers to bank account",
    "Transferencias con error a cuenta bancaria":"Failed transfers to bank account",
    "Cargos a tarjeta de crédito y otros tipos de recuperación de deuda":"Charges to credit card and other debt recovery",
    "Impuestos de producto, envío y envoltura para regalo retenidos":"Product, shipping and gift wrap taxes collected",
    "Impuestos de producto, envío y envoltura para regalo reembolsados":"Product, shipping and gift wrap taxes refunded",
    "Impuesto al valor agregado (IVA) obligatorio de Amazon retenido":"Amazon obligated tax withheld",
    "Ajuste del IVA del producto":"Product VAT adjustment",
    "Impuesto sobre la Renta retenido":"Income tax withheld",
    "subtotales":"subtotals",
}

# ── 西班牙语(ES/Amazon.es) → 英语 ──
ES_TO_EN = {
    "Ingresos":"Income","Costes":"Expenses","Impuesto":"Tax","Transferencias":"Transfers",
    "Venta de productos gestionados por el vendedor":"Seller fulfilled product sales",
    "Reembolso de ventas de productos gestionados por el vendedor":"Seller fulfilled product sale refunds",
    "Venta de productos de Logística de Amazon":"FBA product sales",
    "Reembolso de ventas de productos de Logística de Amazon":"FBA product sale refunds",
    "Crédito de inventario FBA":"FBA inventory credit",
    "Ingresos por liquidación":"FBA liquidation proceeds",
    "Ajustes de ingresos por liquidación de Logística de Amazon":"FBA Liquidations proceeds adjustments",
    "Abonos de envío":"Shipping credits",
    "Reembolso de abonos de envío":"Shipping credit refunds",
    "Abonos de envoltorio para regalo":"Gift wrap credits",
    "Reembolso de abonos por envoltorio para regalo":"Gift wrap credit refunds",
    "Devoluciones promocionales":"Promotional rebates",
    "Reembolso de devoluciones promocionales":"Promotional rebate refunds",
    "Reclamaciones bajo la Garantía de la A a la Z":"A-to-z Guarantee claims",
    "Reversiones de cargo":"Chargebacks",
    "Reembolso de envío de Amazon":"Amazon Shipping Reimbursement Adjustments",
    "IVA de inventario combinado":"Commingling VAT",
    "Reembolso de SAFE-T":"SAFE-T reimbursement",
    "Reversiones de cuentas por cobrar":"Receivables Reversals",
    "Gastos de envío de Amazon":"Amazon Shipping Charge Adjustments",
    "Deducciones de cuentas por cobrar":"Receivables Deductions",
    "Tarifas de venta de pedidos gestionados por el vendedor":"Seller fulfilled selling fees",
    "Tarifas de venta de Logística de Amazon":"FBA selling fees",
    "Reembolso de tarifas de venta":"Selling fee refunds",
    "Tarifas de transacciones de Logística de Amazon":"FBA transaction fees",
    "Tarifas de reembolso de transacciones de Logística de Amazon":"FBA transaction fee refunds",
    "Tarifas de otras transacciones":"Other transaction fees",
    "Reembolso de tarifas de otras transacciones":"Other transaction fee refunds",
    "Tarifas de inventario y de servicios de Logística de Amazon":"FBA inventory and inbound services fees",
    "Compra de etiquetas de envío":"Shipping label purchases",
    "Reembolso de etiquetas de envío":"Shipping label refunds",
    "Ajustes en etiqueta de envío del transportista":"Carrier shipping label adjustments",
    "Tarifas de servicio":"Service fees",
    "Tarifas de administración de reembolsos":"Refund administration fees",
    "Ajustes":"Adjustments",
    "Gastos de publicidad":"Cost of Advertising",
    "Reembolso para anunciante":"Refund for Advertiser",
    "Tarifas del programa de liquidación":"Liquidations fees",
    "Transferencias bancarias":"Transfers to bank account",
    "Transferencias bancarias fallidas":"Failed transfers to bank account",
    "Cargos en tarjetas de crédito y otras recuperaciones de deudas":"Charges to credit card and other debt recovery",
    "Impuestos cobrados por producto, envío y envoltorio para regalo":"Product, shipping and gift wrap taxes collected",
    "Impuestos reembolsados por producto, envío y envoltorio para regalo":"Product, shipping and gift wrap taxes refunded",
    "Impuesto retenido por Amazon":"Amazon obligated tax withheld",
    "subtotal":"subtotals",
}

# ── 英国英语变体 → 标准英语 ──
UK_TO_EN = {
    "Seller-fulfilled product sales":"Seller fulfilled product sales",
    "Seller-fulfilled product sale refunds":"Seller fulfilled product sale refunds",
    "Postage credits":"Shipping credits",
    "Delivery credit refunds":"Shipping credit refunds",
    "Delivery credits":"Shipping credits",
    "Seller-fulfilled selling fees":"Seller fulfilled selling fees",
    "Delivery label purchases":"Shipping label purchases",
    "Delivery label refunds":"Shipping label refunds",
    "Carrier delivery label adjustments":"Carrier shipping label adjustments",
    "Product, delivery and gift wrap taxes collected":"Product, shipping and gift wrap taxes collected",
    "Product, delivery and gift wrap taxes refunded":"Product, shipping and gift wrap taxes refunded",
    "Amazon Shipping Reimbursement":"Amazon Shipping Reimbursement Adjustments",
    "Amazon shipping reimbursement":"Amazon Shipping Reimbursement Adjustments",
    "A-to-z Guarantee Claims":"A-to-z Guarantee claims",
    "Amazon Shipping Charges":"Amazon Shipping Charge Adjustments",
    "Fulfilment by Amazon inventory credit":"FBA inventory credit",
    "Fulfilment by Amazon product sales":"FBA product sales",
    "Product sale refunds(non-FBA)":"Product sale refunds (non-FBA)",
    "Amazon-obligated tax withheld":"Amazon obligated tax withheld",
}

# ── 法语(FR/BE) → 英语 ──
FR_TO_EN = {
    "Revenus":"Income","Dépenses":"Expenses","Taxe":"Tax","Transferts":"Transfers",
    "Vente de produits expédiés par le vendeur":"Seller fulfilled product sales",
    "Remboursement des produits vendus et expédiés par le vendeur":"Seller fulfilled product sale refunds",
    "Vente des produits Expédié par Amazon":"FBA product sales",
    "Remboursement des produits vendus et « Expédié par Amazon »":"FBA product sale refunds",
    "Crédit inventaire FBA":"FBA inventory credit",
    "Produits de liquidation Expédié par Amazon":"FBA liquidation proceeds",
    "Ajustements des recettes des liquidations Expédié par Amazon":"FBA Liquidations proceeds adjustments",
    "Crédits pour l'expédition":"Shipping credits",
    "Remboursement des crédits sur l'expédition":"Shipping credit refunds",
    "Crédits des frais d'emballage":"Gift wrap credits",
    "Remboursement des crédits d'emballage":"Gift wrap credit refunds",
    "Rabais promotionnels":"Promotional rebates",
    "Remboursement des rabais promotionnels":"Promotional rebate refunds",
    "Garanties A à Z":"A-to-z Guarantee claims",
    "Contestations de prélèvement":"Chargebacks",
    "Remboursement des frais d'expédition d'Amazon":"Amazon Shipping Reimbursement Adjustments",
    "Commingling VAT":"Commingling VAT",
    "Remboursement SAFE-T":"SAFE-T reimbursement",
    "Annulations de créances":"Receivables Reversals",
    "Frais d'expédition Amazon":"Amazon Shipping Charge Adjustments",
    "Déductions pour créances":"Receivables Deductions",
    "Frais de vente expédié par le vendeur":"Seller fulfilled selling fees",
    "Frais de vente « Expédié par Amazon »":"FBA selling fees",
    "Remboursement des frais de vente":"Selling fee refunds",
    "Frais de transaction « Expédié par Amazon »":"FBA transaction fees",
    "Remboursement des frais de transaction « Expédié par Amazon »":"FBA transaction fee refunds",
    "Autres frais de transaction":"Other transaction fees",
    "Remboursement d'autres frais de transaction":"Other transaction fee refunds",
    "Frais de stockage et de services logistiques Expédié par Amazon":"FBA inventory and inbound services fees",
    "Achats d'étiquettes d'expédition":"Shipping label purchases",
    "Remboursement des étiquettes d'expédition":"Shipping label refunds",
    "Ajustements de l'étiquette d'expédition du transporteur":"Carrier shipping label adjustments",
    "Frais sur les services":"Service fees",
    "Remboursement des frais administratifs":"Refund administration fees",
    "Ajustements":"Adjustments",
    "Prix de la publicité":"Cost of Advertising",
    "Remboursement pour le publicitaire":"Refund for Advertiser",
    "Frais de liquidation":"Liquidations fees",
    "Transfert sur le compte bancaire":"Transfers to bank account",
    "Versements vers le compte bancaire ayant échoué":"Failed transfers to bank account",
    "Factures de cartes de crédit et autres recouvrements de dettes":"Charges to credit card and other debt recovery",
    "Taxes prélevées sur les produits, les frais d'expédition et l'emballage cadeau":"Product, shipping and gift wrap taxes collected",
    "Taxes remboursées sur les produits, les frais d'expédition et l'emballage cadeau":"Product, shipping and gift wrap taxes refunded",
    "TVA Amazon retenue à la source":"Amazon obligated tax withheld",
    "TVA applicable au Stock sans étiquette":"Commingling VAT",
    "Achats d’étiquettes d’expédition":"Shipping label purchases",
    "Ajustements de l’étiquette d’expédition du transporteur":"Carrier shipping label adjustments",
    "Ajustements des recettes Liquidations Expédié par Amazon":"FBA Liquidations proceeds adjustments",
    "Coût de la publicité":"Cost of Advertising",
    "Crédit du stock Expédié par Amazon":"FBA inventory credit",
    "Crédits d’expédition":"Shipping credits",
    "Crédits pour l’emballage cadeau":"Gift wrap credits",
    "Dédommagement Amazon Shipping":"Amazon Shipping Reimbursement Adjustments",
    "Dédommagement SAFE-T":"SAFE-T reimbursement",
    "Déductions sur créances":"Receivables Deductions",
    "Frais de liquidations":"Liquidations fees",
    "Frais de recouvrement des cartes de crédit et autres créances":"Charges to credit card and other debt recovery",
    "Frais de service":"Service fees",
    "Frais de transaction Expédié par Amazon":"FBA transaction fees",
    "Frais de vente  « Expédié par Amazon »":"FBA selling fees",
    "Frais de vente « Expédié par Amazon »":"FBA selling fees",
    "Frais de vente expédiée par le vendeur":"Seller fulfilled selling fees",
    "Frais de ventes Expédié par Amazon":"FBA selling fees",
    "Remboursement pour l’annonceur":"Refund for Advertiser",
    "Remboursements dans le cadre de la vente de produits (non expédiés par Amazon)":"Seller fulfilled product sale refunds",
    "Remboursements de remises promotionnelles":"Promotional rebate refunds",
    "Remboursements des autres frais de transaction":"Other transaction fee refunds",
    "Remboursements des crédits d’expédition":"Shipping credit refunds",
    "Remboursements des frais de transaction Expédié par Amazon":"FBA transaction fee refunds",
    "Remboursements des frais de vente":"Selling fee refunds",
    "Remboursements des ventes de produits Expédié par Amazon":"FBA product sale refunds",
    "Remboursements des étiquettes d’expédition":"Shipping label refunds",
    "Remboursements du crédit pour l’emballage cadeau":"Gift wrap credit refunds",
    "Retenue de taxes Amazon":"Amazon obligated tax withheld",
    "Retenues sur les remboursements":"Refund administration fees",
    "Réclamations au titre de la Garantie A à Z":"A-to-z Guarantee claims",
    "Rétrofacturations":"Chargebacks",
    "Taxes remboursées concernant les produits, l’expédition et les emballages cadeaux":"Product, shipping and gift wrap taxes refunded",
    "Taxes sur les produits, l’expédition et les emballages cadeaux collectées":"Product, shipping and gift wrap taxes collected",
    "Total des réductions":"Promotional rebates",
    "Transferts sur le compte bancaire":"Transfers to bank account",
    "Ventes de produits (non expédiés par Amazon)":"Seller fulfilled product sales",
    "Ventes de produits Expédié par Amazon":"FBA product sales",
    "Échec des transferts sur le compte bancaire":"Failed transfers to bank account",
    "sous-totaux":"subtotals",
}

# ── 意大利语(IT) → 英语 ──
IT_TO_EN = {
    "Ricavi":"Income","Spese":"Expenses","Imposte":"Tax","Trasferimenti":"Transfers",
    "Vendite articoli gestiti dal venditore":"Seller fulfilled product sales",
    "Rimborsi per vendita articoli gestiti dal venditore":"Seller fulfilled product sale refunds",
    "Vendite articoli gestiti con Logistica di Amazon":"FBA product sales",
    "Rimborsi per articoli gestiti con Logistica di Amazon":"FBA product sale refunds",
    "Credito dell'inventario FBA":"FBA inventory credit",
    "Ricavi della liquidazione di Logistica di Amazon":"FBA liquidation proceeds",
    "Aggiustamenti sui ricavi di liquidazione di Logistica di Amazon":"FBA Liquidations proceeds adjustments",
    "Accrediti per le spedizioni":"Shipping credits",
    "Rimborsi per accrediti spedizioni":"Shipping credit refunds",
    "Accrediti per confezione regalo":"Gift wrap credits",
    "Rimborsi per accrediti confezioni regalo":"Gift wrap credit refunds",
    "Sconti promozionali":"Promotional rebates",
    "Rimborsi per sconti promozionali":"Promotional rebate refunds",
    "Reclami dalla A alla Z":"A-to-z Guarantee claims",
    "Chargeback":"Chargebacks",
    "Rimborso spedizione Amazon":"Amazon Shipping Reimbursement Adjustments",
    "Commingling VAT":"Commingling VAT",
    "Rimborso SAFE-T (Seller Assurance for E-commerce Transactions)":"SAFE-T reimbursement",
    "Storni dei crediti":"Receivables Reversals",
    "Costi di spedizione Amazon":"Amazon Shipping Charge Adjustments",
    "Detrazioni dei crediti":"Receivables Deductions",
    "Commissioni di vendita gestita dal venditore":"Seller fulfilled selling fees",
    "Commissioni di vendita con Logistica di Amazon":"FBA selling fees",
    "Rimborsi per commissioni di vendita":"Selling fee refunds",
    "Commissioni Logistica di Amazon per transazione":"FBA transaction fees",
    "Rimborsi commissioni Logistica di Amazon per transazione":"FBA transaction fee refunds",
    "Altri costi relativi alle transazioni":"Other transaction fees",
    "Rimborsi per altri costi relativi alle transazioni":"Other transaction fee refunds",
    "Costi di gestione dell'inventario e per i servizi di Logistica di Amazon":"FBA inventory and inbound services fees",
    "Acquisti delle etichette di spedizione":"Shipping label purchases",
    "Rimborsi relativi alle etichette di spedizione":"Shipping label refunds",
    "Modifiche etichetta di spedizione corriere":"Carrier shipping label adjustments",
    "Commissioni di servizio":"Service fees",
    "Costi amministrativi relativi ai rimborsi":"Refund administration fees",
    "Modifiche":"Adjustments",
    "Costo della pubblicità":"Cost of Advertising",
    "Rimborso per inserzionista":"Refund for Advertiser",
    "Tariffe del Programma di liquidazione":"Liquidations fees",
    "Trasferimenti sul conto corrente":"Transfers to bank account",
    "Trasferimenti non riusciti sul conto corrente":"Failed transfers to bank account",
    "Costi relativi alla carta di credito e ad altri tipi di recupero crediti":"Charges to credit card and other debt recovery",
    "Imposte relative a prodotto, spedizione e confezione regalo riscosse":"Product, shipping and gift wrap taxes collected",
    "Imposte relative a prodotto, spedizione e confezione regalo rimborsate":"Product, shipping and gift wrap taxes refunded",
    "Ritenuta fiscale obbligatoria da Amazon":"Amazon obligated tax withheld",
    "Totale parziale":"subtotals",
}

# ── 波兰语(PL) → 英语 ──
PL_TO_EN = {
    "Przychód":"Income","Wydatki":"Expenses","Podatek":"Tax","Przelewy":"Transfers",
    "Sprzeda produktów (inna ni FBA)":"Seller fulfilled product sales",
    "Zwroty kosztów sprzeday produktów (innej ni FBA)":"Seller fulfilled product sale refunds",
    "FBA — sprzeda produktów":"FBA product sales",
    "FBA — zwroty kosztów sprzeday produktów":"FBA product sale refunds",
    "FBA — rodki na pokrycie zapasów":"FBA inventory credit",
    "FBA — przychody z likwidacji":"FBA liquidation proceeds",
    "FBA Liquidations proceeds adjustments":"FBA Liquidations proceeds adjustments",
    "Noty kredytowe za wysyłk":"Shipping credits",
    "Zwroty do not kredytowych za wysyłk":"Shipping credit refunds",
    "rodki na pokrycie pakowania na prezent":"Gift wrap credits",
    "Zwroty rodków na pokrycie pakowania na prezent":"Gift wrap credit refunds",
    "Rabaty promocyjne":"Promotional rebates",
    "Zwroty kosztów rabatów promocyjnych":"Promotional rebate refunds",
    "Roszczenia w ramach gwarancji od A do Z":"A-to-z Guarantee claims",
    "Obcienia zwrotne":"Chargebacks",
    "Amazon Shipping Reimbursement Adjustments":"Amazon Shipping Reimbursement Adjustments",
    "VAT za wspólne zapasy":"Commingling VAT",
    "Zwrot kosztów SAFE-T":"SAFE-T reimbursement",
    "Odwrócenie nalenoci":"Receivables Reversals",
    "Opłaty za wysyłk Amazon":"Amazon Shipping Charge Adjustments",
    "Potrcenia nalenoci":"Receivables Deductions",
    "Opłaty za sprzeda realizowan przez Sprzedawc":"Seller fulfilled selling fees",
    "FBA — opłaty za sprzeda":"FBA selling fees",
    "Zwroty opłat za sprzeda":"Selling fee refunds",
    "FBA — opłaty transakcyjne":"FBA transaction fees",
    "FBA — zwroty opłat transakcyjnych":"FBA transaction fee refunds",
    "Inne opłaty transakcyjne":"Other transaction fees",
    "Zwroty innych opłat transakcyjnych":"Other transaction fee refunds",
    "FBA — opłaty za zapasy i usługi zwizane z wysyłk":"FBA inventory and inbound services fees",
    "Zakupy etykiet wysyłkowych":"Shipping label purchases",
    "Zwroty kosztów etykiet wysyłkowych":"Shipping label refunds",
    "Korekty do etykiet wysyłkowych przez przewonika":"Carrier shipping label adjustments",
    "Opłaty za usługi":"Service fees",
    "Opłata manipulacyjna za zwrot kosztów":"Refund administration fees",
    "Korekty":"Adjustments",
    "Koszt reklamy":"Cost of Advertising",
    "Zwrot kosztów dla reklamodawcy":"Refund for Advertiser",
    "Opłaty za likwidacj":"Liquidations fees",
    "Przelewy na rachunek bankowy":"Transfers to bank account",
    "Nieudane przelewy na rachunek bankowy":"Failed transfers to bank account",
    "Opłaty ztytułu kart kredytowych iinnych form odzyskiwania nalenoci":"Charges to credit card and other debt recovery",
    "Potrcone podatki od produktów, wysyłki i pakowania na prezent":"Product, shipping and gift wrap taxes collected",
    "Potrcone podatki od produktów, wysyłki i pakowania":"Product, shipping and gift wrap taxes collected",
    "Zwrócone podatki od produktów, wysyłki i pakowania":"Product, shipping and gift wrap taxes refunded",
    "Podatek potrcony przez Amazon (w ramach zobowiza)":"Amazon obligated tax withheld",
    "sumy czciowe":"subtotals",
}

ALL_LANG_MAPS = [DE_TO_EN, SE_TO_EN, NL_TO_EN, JP_TO_EN, MX_TO_EN, UK_TO_EN, FR_TO_EN, IT_TO_EN, PL_TO_EN, ES_TO_EN]

def normalize_en(f):
    f = f.strip()
    for m in ALL_LANG_MAPS:
        if f in m:
            return m[f]
    return f

def get_cn(f):
    en = normalize_en(f)
    return FIELD_CN.get(en, en)

def parse_num(s):
    if not s: return None
    s = str(s).strip()
    # 统一负号（Unicode减号 U+2212 → ASCII减号）
    s = s.replace("\u2212", "-").replace("\u2013", "-")
    if not s or s in ("-","–","−"): return None
    # 欧洲格式1：1.234,56（德/荷/波兰）
    if re.match(r"^-?[\d\.]+,\d{1,2}$", s.replace(" ","")):
        s = s.replace(".","").replace(",",".")
    # 欧洲格式2：77664,85（瑞典，无千分位点）
    elif re.match(r"^-?\d+,\d{1,2}$", s.replace(" ","")):
        s = s.replace(",",".")
    else:
        s = s.replace(",","")
    try: return float(s)
    except: return None

def is_num_token(t):
    # 支持 ASCII 减号和 Unicode 减号 (U+2212)
    c = t.replace(",","").replace(".","").lstrip("-").lstrip("\u2212").lstrip("\u2013")
    return bool(c) and c.isdigit()

def group_col_rows(wlist, tol=2.5):
    rows = {}
    for w in wlist:
        k = round(w["top"]/tol)*tol
        rows.setdefault(k,[]).append(w)
    return [(k, sorted(v, key=lambda x:x["x0"])) for k,v in sorted(rows.items())]

def merge_orphan_nums(rows):
    """合并只有数字的行到相邻的有文字的行（前后都检查）"""
    merged = []
    i = 0
    while i < len(rows):
        top, wlist = rows[i]
        texts = [w for w in wlist if not is_num_token(w["text"])]
        nums = [w for w in wlist if is_num_token(w["text"])]
        if not texts and nums:
            # 尝试合并到下一行（有文字）
            if i+1 < len(rows):
                next_top, next_wlist = rows[i+1]
                if next_top - top < 6:
                    next_texts = [w for w in next_wlist if not is_num_token(w["text"])]
                    if next_texts:
                        merged.append((top, next_wlist + wlist))
                        i += 2
                        continue
            # 尝试合并到前一行（有文字）
            if merged:
                prev_top, prev_wlist = merged[-1]
                if top - prev_top < 6:
                    prev_texts = [w for w in prev_wlist if not is_num_token(w["text"])]
                    if prev_texts:
                        merged[-1] = (prev_top, prev_wlist + wlist)
                        i += 1
                        continue
        merged.append((top, wlist))
        i += 1
    return merged

def parse_column(col_rows, debit_x, credit_x):
    """
    解析单列的行数据。
    debit_x, credit_x: 该列的 Debits/Credits 表头 x 坐标
    返回: (details_list, subtotals_dict)
    details_list: [(section, field_original, field_en, debit, credit)]
    subtotals_dict: {section: (debit_total, credit_total)}
    """
    details, subtotals = [], {}
    current_section = None
    TOL = 60  # 数值归属容差

    def classify(num_x):
        dd = abs(num_x - debit_x)
        dc = abs(num_x - credit_x)
        return "debit" if dd <= dc else "credit"

    for top, wlist in col_rows:
        text_words = [w for w in wlist if not is_num_token(w["text"])]
        num_words = [(w["x0"], w["text"]) for w in wlist if is_num_token(w["text"])]
        field_text = " ".join(w["text"] for w in text_words).strip()
        ft_lower = field_text.lower().strip()

        # 跳过无意义行
        if any(ft_lower.startswith(p) for p in SKIP_STARTS): continue
        if not field_text and not num_words: continue
        # 段落标题：单词是大类名 + 一个数值
        words_in_field = field_text.split()
        if (len(words_in_field) == 1 and
            words_in_field[0].lower() in ALL_SECTIONS and
            len(num_words) == 1):
            sec_en = normalize_en(words_in_field[0])
            current_section = sec_en.capitalize()
            continue

        # 跳过表头行。德语的收入章节名和 Credits 列头都叫 Einnahmen，
        # 因此必须在章节识别之后再过滤表头。
        if re.match(r"^(debits|credits|belastungen|einnahmen)(\s+(debits|credits|belastungen|einnahmen))?$", ft_lower): continue

        # subtotals
        if ft_lower in SUBTOTAL_WORDS:
            if current_section and num_words:
                sorted_nums = sorted(num_words, key=lambda x: x[0])
                aligned_nums = [n for n in sorted_nums if n[0] >= min(debit_x, credit_x) - 25]
                if aligned_nums:
                    sorted_nums = aligned_nums
                if len(sorted_nums) >= 2:
                    debit_vals, credit_vals = [], []
                    for nx, nt in sorted_nums:
                        v = parse_num(nt)
                        if v is None:
                            continue
                        if classify(nx) == "debit":
                            debit_vals.append(v)
                        else:
                            credit_vals.append(v)
                    d = round(sum(debit_vals), 2) if debit_vals else None
                    c = round(sum(credit_vals), 2) if credit_vals else None
                    subtotals[current_section] = (d, c)
                elif len(sorted_nums) == 1:
                    v = parse_num(sorted_nums[0][1])
                    if v is not None and v < 0:
                        subtotals[current_section] = (v, None)
                    else:
                        subtotals[current_section] = (None, v)
            continue

        # 明细行
        if current_section and field_text:
            field_en = normalize_en(field_text)
            if field_en.lower() in HEADER_WORDS: continue
            if not field_en or field_en.lower() in SUBTOTAL_WORDS: continue

            debit, credit = None, None
            for nx, nt in num_words:
                v = parse_num(nt)
                if v is None: continue
                dc = classify(nx)
                if dc == "debit":
                    debit = v
                else:
                    credit = v

            # 只有一个数值且为0时，根据字段类型判断归属
            if len(num_words) == 1 and debit is None and credit is None:
                v = parse_num(num_words[0][1])
                if v == 0.0:
                    # 0值：用 x 坐标判断
                    nx = num_words[0][0]
                    dc = classify(nx)
                    if dc == "debit":
                        debit = 0.0
                    else:
                        credit = 0.0

            details.append((current_section, field_text, field_en, debit, credit))

    return details, subtotals


def extract_positioned_text(pdf_path):
    """Return text fragments with pdfplumber-like x0/top keys.

    pdfplumber is preferred when available. The pypdf fallback uses the PDF text
    transform matrix and intentionally keeps full line fragments, which matches
    these Amazon statement PDFs well and preserves localized field labels.
    """
    if pdfplumber is not None:
        with pdfplumber.open(pdf_path) as pdf:
            all_words = []
            for page in pdf.pages:
                all_words.extend(page.extract_words(x_tolerance=3, y_tolerance=3))
            return all_words

    reader = PdfReader(pdf_path)
    words = []
    page_offset = 0.0
    for page in reader.pages:
        height = float(page.mediabox.height)

        def visitor(text, cm, tm, font_dict, font_size):
            value = (text or "").strip()
            if not value:
                return
            # Ignore pypdf's duplicate low-level text cache entries. Real content
            # in these PDFs carries a non-zero current transformation matrix.
            if abs(cm[4]) < 1e-6 and abs(cm[5]) < 1e-6:
                return
            words.append({
                "text": value,
                "x0": float(cm[4]),
                "top": page_offset + height - float(cm[5]),
            })

        page.extract_text(visitor_text=visitor)
        page_offset += height
    return words


def quarter_from_month(month):
    return f"Q{((month - 1) // 3) + 1}" if month else None


def parse_filename_period(fname):
    m = re.search(r"(?<!\d)(\d{4})(\d{2})月", fname)
    if m:
        year = int(m.group(1))
        month = int(m.group(2))
        if 1 <= month <= 12:
            return year, month, quarter_from_month(month)
        return None, None, None
    m = re.search(r"(?<!\d)(\d{4})(0[1-9]|1[0-2])(?=[-_])", fname)
    if m:
        year = int(m.group(1))
        month = int(m.group(2))
        return year, month, quarter_from_month(month)
    m = re.search(r"(\d{4})Q([1-4])-Q([1-4])", fname, re.IGNORECASE)
    if m:
        return int(m.group(1)), None, f"Q{m.group(2)}-Q{m.group(3)}"
    m = re.search(r"(\d{4})Q([1-4])", fname, re.IGNORECASE)
    if m:
        return int(m.group(1)), None, f"Q{m.group(2)}"
    return None, None, None


def parse_period_bounds(period_text):
    matches = re.findall(r"(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d+,\s+(\d{4})", period_text or "")
    if not matches:
        return None, None, None, None, None, None
    start_month = MONTH_ABBR_TO_NUM.get(matches[0][0])
    start_year = int(matches[0][1])
    end_month = MONTH_ABBR_TO_NUM.get(matches[-1][0])
    end_year = int(matches[-1][1])
    return (
        start_year,
        start_month,
        quarter_from_month(start_month),
        end_year,
        end_month,
        quarter_from_month(end_month),
    )


def parse_period_start(period_text):
    py, pm, pq, _ey, _em, _eq = parse_period_bounds(period_text)
    return py, pm, pq


def period_quarter_range(start_year, start_quarter, end_year, end_quarter):
    if not (start_year and start_quarter and end_year and end_quarter):
        return None
    if start_year != end_year:
        return f"{start_year}{start_quarter}-{end_year}{end_quarter}"
    if start_quarter == end_quarter:
        return start_quarter
    return f"{start_quarter}-{end_quarter}"


def derive_report_period(fname, period_text):
    fy, fm, fq = parse_filename_period(fname)
    py, pm, pq, pey, pem, peq = parse_period_bounds(period_text)
    pqr = period_quarter_range(py, pq, pey, peq)

    if py and pm:
        year, month, quarter = py, pm, pqr or pq
        if fy and fm:
            if (fy, fm) == (py, pm):
                status = "✓ 文件名与报告期一致"
                note = ""
            else:
                status = "文件名年月与报告期不一致，已采用报告期"
                note = f"文件名={fy}-{fm:02d}; 报告期={py}-{pm:02d}"
        elif fy and fq and "-" in fq:
            if (fy, fq) == (py, pqr):
                status = "✓ 文件名季度范围与报告期一致"
                note = ""
            else:
                status = "文件名季度范围与报告期不一致，已采用报告期"
                note = f"文件名={fy}{fq}; 报告期={py}{pqr}"
        elif fy and fq:
            if (fy, fq) == (py, pq):
                status = "✓ 文件名季度与报告期一致"
                note = ""
            else:
                status = "文件名季度与报告期不一致，已采用报告期"
                note = f"文件名={fy}{fq}; 报告期={py}{pq}"
        else:
            status = "文件名未匹配标准年月，已采用报告期"
            note = f"文件名未识别出 YYYYMM月；报告期={py}-{pm:02d}"
    elif fy and fm:
        year, month, quarter = fy, fm, fq
        status = "未解析到报告期，已采用文件名年月"
        note = f"文件名={fy}-{fm:02d}"
    elif fy and fq:
        year, month, quarter = fy, None, fq
        status = "未解析到报告期，已采用文件名季度"
        note = f"文件名={fy}{fq}"
    else:
        year, month, quarter = None, None, None
        status = "未解析到文件名年月或报告期年月"
        note = ""

    return {
        "year": year,
        "month": month,
        "quarter": quarter,
        "filename_year": fy,
        "filename_month": fm,
        "filename_quarter": fq,
        "period_year": py,
        "period_month": pm,
        "period_quarter": pq,
        "filename_audit_status": status,
        "filename_audit_note": note,
    }


def parse_file_period(fname, period_text):
    meta = derive_report_period(fname, period_text)
    return meta["year"], meta["month"], meta["quarter"]


MONTH_ABBR_TO_NUM = {"Jan":1,"Feb":2,"Mar":3,"Apr":4,"May":5,"Jun":6,
                     "Jul":7,"Aug":8,"Sep":9,"Oct":10,"Nov":11,"Dec":12}


def extract_pdf(pdf_path, store, country, country_code=None):
    meta = {"store":store,"country":country,"country_code":country_code or country,
            "source_file":os.path.basename(pdf_path),"source_path":str(pdf_path),
            "currency":"","period":"",
            "year":None,"month":None,"quarter":None,
            "filename_year":None,"filename_month":None,"filename_quarter":None,
            "period_year":None,"period_month":None,"period_quarter":None,
            "filename_audit_status":"","filename_audit_note":"",
            "display_name":"","legal_name":""}
    summaries, details, errors, checks = [], [], [], []

    all_words = sorted(extract_positioned_text(pdf_path), key=lambda w: (w["top"], w["x0"]))
    full_text = " ".join(w["text"] for w in all_words)

    # 货币（多语言）
    m = re.search(r"All amounts in ([A-Z]{3})", full_text)
    if m:
        meta["currency"] = m.group(1)
    elif re.search(r"Todos los importes en ([A-Z]{3})", full_text):
        meta["currency"] = re.search(r"Todos los importes en ([A-Z]{3})", full_text).group(1)
    elif re.search(r"Alle bedragen in ([A-Z]{3})", full_text):
        meta["currency"] = re.search(r"Alle bedragen in ([A-Z]{3})", full_text).group(1)
    elif re.search(r"Wszystkie kwoty w ([A-Z]{3})", full_text):
        meta["currency"] = re.search(r"Wszystkie kwoty w ([A-Z]{3})", full_text).group(1)
    elif re.search(r"Tous les montants sont en ([A-Z]{3})", full_text):
        meta["currency"] = re.search(r"Tous les montants sont en ([A-Z]{3})", full_text).group(1)
    elif re.search(r"Tutti gli importi sono espressi in ([A-Z]{3})", full_text):
        meta["currency"] = re.search(r"Tutti gli importi sono espressi in ([A-Z]{3})", full_text).group(1)
    elif re.search(r"Wszystkie kwoty w ([A-Z]{3})", full_text):
        meta["currency"] = re.search(r"Wszystkie kwoty w ([A-Z]{3})", full_text).group(1)
    elif "Euro" in full_text or "EUR" in full_text: meta["currency"] = "EUR"
    elif "Alla belopp i kr" in full_text: meta["currency"] = "SEK"
    elif "単位は円" in full_text or "指定のない場合、単位は円" in full_text: meta["currency"] = "JPY"
    elif "importes en MXN" in full_text: meta["currency"] = "MXN"
    elif "amounts in local" in full_text:
        # SA=SAR, AE=AED
        if meta["country_code"] == "SA": meta["currency"] = "SAR"
        elif meta["country_code"] == "AE": meta["currency"] = "AED"

    # 报告期（多语言）
    period_patterns = [
        r"Account activity from (.+?) through (.+?)(?:All amounts|$)",   # EN (US/CA/AU/SA/AE)
        r"Account activity from (.+?) to (.+?)(?:All amounts|$)",        # EN-UK
        r"Vorgänge vom (.+?) bis (.+?)(?:Alle|$)",                       # DE
        r"Kontoaktivitet från och med (.+?) till och med (.+?)(?:Alla|$)",# SE
        r"Accountactiviteit van (.+?) tot en met (.+?)(?:Alle|$)",       # NL
        r"Aktywno.{0,5} na koncie od (.+?) do (.+?)(?:Wszystkie|$)",     # PL
        r"Actividad de la cuenta desde (.+?) hasta (.+?)(?:Todos|$)",    # MX/ES
        r"(.+?) から (.+?) までのアカウント履歴",                          # JP
        r"Attività dell.account dal (.+?) al (.+?)(?:Tutti|$)",          # IT
        r"Activité du compte du (.+?) au (.+?)(?:Tous|$)",               # FR
        r"Activité du compte de (.+?) à (.+?)(?:Tous|$)",               # BE(FR)
        r"Actividad de la cuenta desde el (.+?) hasta el (.+?)(?:Todos|$)", # ES(Amazon.es)
        r"Attività dell.account da (.+?) a (.+?)(?:Tutti|$)",           # IT
        r"Aktywno.{0,5} na koncie od (.+?) do (.+?)(?:Wszystkie|$)",    # PL
        r"Actividad de la cuenta del (.+?) al (.+?)(?:Todos|$)",         # ES
    ]
    period_line_patterns = [
        r"Account activity from (.+?) through (.+)$",
        r"Account activity from (.+?) to (.+)$",
        r"Vorgänge vom (.+?) bis (.+)$",
        r"Kontoaktivitet från och med (.+?) till och med (.+)$",
        r"Accountactiviteit van (.+?) tot en met (.+)$",
        r"Aktywno.{0,5} na koncie od (.+?) do (.+)$",
        r"Actividad de la cuenta desde (.+?) hasta (.+)$",
        r"(.+?) から (.+?) までのアカウント履歴$",
        r"Attività dell.account dal (.+?) al (.+)$",
        r"Attività dell.account da (.+?) a (.+)$",
        r"Activité du compte du (.+?) au (.+)$",
        r"Activité du compte de (.+?) à (.+)$",
        r"Actividad de la cuenta desde el (.+?) hasta el (.+)$",
        r"Actividad de la cuenta del (.+?) al (.+)$",
    ]
    for txt in [w["text"] for w in all_words]:
        for pat in period_line_patterns:
            m = re.search(pat, txt)
            if m:
                meta["period"] = f"{m.group(1).strip()} ~ {m.group(2).strip()}"
                break
        if meta["period"]:
            break

    for pat in period_patterns:
        if meta["period"]:
            break
        m = re.search(pat, full_text)
        if m:
            p1 = m.group(1).strip()
            p2 = m.group(2).strip()
            # 清理：只保留日期部分（去掉前面可能混入的非日期文字）
            # 日期格式：Jan 1, 2026 或 2026年1月1日 等
            date_m = re.search(r"(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec).+", p1)
            if date_m: p1 = date_m.group(0)
            date_m = re.search(r"(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec).+", p2)
            if date_m: p2 = date_m.group(0)
            meta["period"] = f"{p1} ~ {p2}"
            break

    meta.update(derive_report_period(os.path.basename(pdf_path), meta["period"]))

    # Display/Legal name
    m = re.search(r"Display name[:\s]+(\S+)", full_text)
    if m: meta["display_name"] = m.group(1)
    m = re.search(r"Anzeigename[:\s]+(\S+)", full_text)
    if m: meta["display_name"] = m.group(1)
    m = re.search(r"Legal name[:\s]+(\S+)", full_text)
    if m: meta["legal_name"] = m.group(1)
    m = re.search(r"Eingetragener Firmenname[:\s]+(\S+)", full_text)
    if m: meta["legal_name"] = m.group(1)

    COL_SPLIT = 395

    # ── Summaries 提取（top 95-170，数值在最右边 x>700）──
    summaries_map = {}
    for w in all_words:
        if not (95 <= w["top"] <= 170): continue
        if w["x0"] >= 100: continue  # 只看左边的字段名
        field_en = normalize_en(w["text"])
        if field_en.lower() not in ALL_SECTIONS: continue
        # 找同一行最右边的数值
        row_nums = [x for x in all_words
                    if abs(x["top"] - w["top"]) < 3 and x["x0"] > 600 and is_num_token(x["text"])]
        if row_nums:
            val = parse_num(sorted(row_nums, key=lambda x: x["x0"])[-1]["text"])
            if val is not None and field_en.capitalize() not in summaries_map:
                summaries_map[field_en.capitalize()] = val
                summaries.append({"section": field_en.capitalize(), "total": val})

    # ── 动态检测 Debits/Credits 列坐标 ──
    # 找 top > 150 的 Debits/Credits 表头
    deb_xs_all, cred_xs_all = [], []
    for w in all_words:
        if w["top"] <= 150: continue
        if w["text"].lower() in ("debits","belastungen"):
            deb_xs_all.append((w["x0"], w["top"]))
        if w["text"].lower() in ("credits","einnahmen") and w["x0"] > 100:
            cred_xs_all.append((w["x0"], w["top"]))

    # 左列（x < COL_SPLIT）和右列（x >= COL_SPLIT）分别取坐标
    left_deb_x = min((x for x,_ in deb_xs_all if x < COL_SPLIT), default=285)
    left_cred_x = min((x for x,_ in cred_xs_all if x < COL_SPLIT), default=353)
    right_deb_x = min((x for x,_ in deb_xs_all if x >= COL_SPLIT), default=669)
    right_cred_x = min((x for x,_ in cred_xs_all if x >= COL_SPLIT), default=737)

    # ── 分左右列，各自分行，合并孤立数字行 ──
    detail_words = [w for w in all_words if w["top"] > 170]
    left_words = [w for w in detail_words if w["x0"] < COL_SPLIT]
    right_words = [w for w in detail_words if w["x0"] >= COL_SPLIT]

    left_rows = merge_orphan_nums(group_col_rows(left_words))
    right_rows = merge_orphan_nums(group_col_rows(right_words))

    # ── 解析各列 ──
    left_details, left_subtotals = parse_column(left_rows, left_deb_x, left_cred_x)
    right_details, right_subtotals = parse_column(right_rows, right_deb_x, right_cred_x)

    all_details = left_details + right_details
    all_subtotals = {**left_subtotals, **right_subtotals}

    for sec, field_original, field_en, d, c in all_details:
        details.append({"section":sec,"field_original":field_original,
                        "field_en":field_en,"debit":d,"credit":c})

    # ── 交叉核验 ──
    sec_map = {}
    for det in details:
        sec_map.setdefault(det["section"],[]).append(det)

    for sec, items in sec_map.items():
        total_d = round(sum(it["debit"] for it in items if it["debit"] is not None), 2)
        total_c = round(sum(it["credit"] for it in items if it["credit"] is not None), 2)
        net = round(total_d + total_c, 2)

        if sec in summaries_map:
            diff = round(abs(net - summaries_map[sec]), 2)
            if diff > 0.05:
                errors.append(f"[{sec}] 明细净值{net} ≠ Summaries合计{summaries_map[sec]}，差异={diff}")

        if sec in all_subtotals:
            st_d, st_c = all_subtotals[sec]
            if st_d is not None and abs(total_d - st_d) > 0.05:
                errors.append(f"[{sec}] 借方小计{total_d} ≠ subtotals借方{st_d}，差异={round(abs(total_d-st_d),2)}")
            if st_c is not None and abs(total_c - st_c) > 0.05:
                errors.append(f"[{sec}] 贷方小计{total_c} ≠ subtotals贷方{st_c}，差异={round(abs(total_c-st_c),2)}")
        st_d, st_c = all_subtotals.get(sec, (None, None))
        summary_total = summaries_map.get(sec)
        checks.append({
            "section": sec,
            "detail_debits": total_d,
            "detail_credits": total_c,
            "detail_net": net,
            "summary_total": summary_total,
            "subtotal_debits": st_d,
            "subtotal_credits": st_c,
            "summary_diff": round(net - summary_total, 2) if summary_total is not None else None,
            "debit_diff": round(total_d - st_d, 2) if st_d is not None else None,
            "credit_diff": round(total_c - st_c, 2) if st_c is not None else None,
        })

    return {"meta":meta,"summaries":summaries,"details":details,"checks":checks,"errors":errors}

def collect_pdfs(base_dir):
    pdfs = []
    for store in sorted(os.listdir(base_dir)):
        sp = os.path.join(base_dir, store)
        if not os.path.isdir(sp) or store.startswith(".") or store in ("script", "outputs"): continue
        for root, dirs, files in os.walk(sp):
            dirs[:] = sorted([d for d in dirs if not d.startswith(".")])
            for fname in sorted(files):
                if not fname.lower().endswith(".pdf"): continue
                if re.search(r"Q[1-4]-Q[1-4]", fname, re.IGNORECASE): continue
                pdf_path = os.path.join(root, fname)
                qname = re.search(r"^(.+?)-([A-Z]{2})-\d{4}Q[1-4](?:-Q[1-4])?\b", fname, re.IGNORECASE)
                m = re.search(r"-([A-Z]{2})-\d{4}Q\d", fname, re.IGNORECASE)
                country_name = Path(root).name
                filename_country = next((name for name in COUNTRY_NAME_TO_CODE if name in fname), None)
                if filename_country:
                    country_name = filename_country
                cc = m.group(1).upper() if m else COUNTRY_NAME_TO_CODE.get(country_name, "XX")
                inferred_store = qname.group(1) if qname else store
                if country_name not in COUNTRY_NAME_TO_CODE and qname:
                    country_name = CODE_TO_COUNTRY_NAME.get(cc, country_name)
                pdfs.append((pdf_path, inferred_store, country_name, cc))
    return pdfs

HEADERS = ["店铺","国家/站点","站点代码","货币","报告期","年份","月份","季度",
           "大类","原始字段","英文字段","中文标准字段",
           "借方(Debits)","贷方(Credits)","净值","核验状态","来源文件"]
CHECK_HEADERS = ["店铺","国家/站点","站点代码","货币","报告期","年份","月份","季度",
                 "大类","明细借方合计","明细贷方合计","明细净值",
                 "Summaries合计","Subtotals借方","Subtotals贷方",
                 "净值差异","借方差异","贷方差异","核验状态","来源文件"]
FILENAME_AUDIT_HEADERS = ["店铺","国家/站点","站点代码","来源文件","文件路径","报告期",
                          "文件名年份","文件名月份","报告期年份","报告期月份",
                          "最终年份","最终月份","最终季度","状态","说明"]
SECTION_BG = {"Summaries":"FFD6E4F0","Income":"FFE2EFDA",
              "Expenses":"FFFCE4D6","Transfers":"FFFFF2CC","Tax":"FFEDEDED"}
HDR_BG, HDR_FG = "FF1F4E79", "FFFFFFFF"

def tborder():
    s = Side(style="thin", color="FFB0B0B0")
    return Border(left=s,right=s,top=s,bottom=s)

def wrow(ws, row, data, bg, bold=False, err=False):
    bd = tborder()
    for ci, val in enumerate(data, 1):
        c = ws.cell(row=row, column=ci, value=val)
        c.font = Font(bold=bold, size=9)
        c.border = bd
        c.alignment = Alignment(vertical="center")
        if ci in (13,14,15) and isinstance(val, (int, float)):
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.number_format = "#,##0.00"
        if ci == 16 and err and val:
            c.fill = PatternFill("solid", fgColor="FFFFC7CE")
            c.font = Font(bold=True, color="FF9C0006", size=9)
        else:
            c.fill = PatternFill("solid", fgColor=bg)

def write_excel(all_results, output_path):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb = openpyxl.Workbook()
    ws_all = wb.active
    ws_all.title = "All"
    sheet_names = ["All", "Summaries", "Income", "Expenses", "Transfers", "Tax", "Checks", "Filename_Audit"]
    sheets = {"All": ws_all}
    for name in sheet_names[1:]:
        sheets[name] = wb.create_sheet(name)

    def setup_sheet(ws, headers):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = Font(bold=True, color=HDR_FG, size=10)
            c.fill = PatternFill("solid", fgColor=HDR_BG)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = tborder()
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for name in sheet_names:
        if name == "Checks":
            headers = CHECK_HEADERS
        elif name == "Filename_Audit":
            headers = FILENAME_AUDIT_HEADERS
        else:
            headers = HEADERS
        setup_sheet(sheets[name], headers)

    next_row = {name: 2 for name in sheet_names}

    def add_data_row(sheet_name, row_data, bg, bold=False, err=False):
        for name in ("All", sheet_name):
            ws = sheets[name]
            wrow(ws, next_row[name], row_data, bg, bold=bold, err=err)
            next_row[name] += 1

    def add_check_row(row_data, err=False):
        ws = sheets["Checks"]
        row = next_row["Checks"]
        bd = tborder()
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.font = Font(bold=(ci == 19 and err), color=("FF9C0006" if ci == 19 and err else "FF000000"), size=9)
            c.border = bd
            c.alignment = Alignment(vertical="center")
            if ci in range(10, 19) and isinstance(val, (int, float)):
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.number_format = "#,##0.00"
            if err and ci == 19:
                c.fill = PatternFill("solid", fgColor="FFFFC7CE")
            else:
                c.fill = PatternFill("solid", fgColor="FFFFFFFF")
        next_row["Checks"] += 1

    def add_filename_audit_row(row_data, err=False):
        ws = sheets["Filename_Audit"]
        row = next_row["Filename_Audit"]
        bd = tborder()
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.font = Font(bold=(ci == 14 and err), color=("FF9C0006" if ci == 14 and err else "FF000000"), size=9)
            c.border = bd
            c.alignment = Alignment(vertical="center")
            if ci in (7, 8, 9, 10, 11, 12):
                c.alignment = Alignment(horizontal="center", vertical="center")
            if err and ci == 14:
                c.fill = PatternFill("solid", fgColor="FFFFC7CE")
            else:
                c.fill = PatternFill("solid", fgColor="FFFFFFFF")
        next_row["Filename_Audit"] += 1

    duplicate_keys = {}
    for res in all_results:
        m = res["meta"]
        if not (m["year"] and m["month"]):
            continue
        key = (m["store"], m["country_code"], m["year"], m["month"], m["quarter"])
        duplicate_keys.setdefault(key, []).append(m["source_file"])

    for res in all_results:
        m = res["meta"]
        base = [m["store"], m["country"], m["country_code"], m["currency"], m["period"],
                m["year"], m["month"], m["quarter"]]
        has_err = bool(res["errors"])
        err_str = "; ".join(res["errors"]) if has_err else "✓ 核验通过"

        for idx, s in enumerate(res["summaries"]):
            sec = s["section"]
            row_data = base + ["Summaries", sec, sec, get_cn(sec),
                               None, None, s["total"],
                               err_str if idx == 0 else "", m["source_file"]]
            add_data_row("Summaries", row_data, SECTION_BG.get("Summaries","FFFFFFFF"),
                         bold=True, err=(has_err and idx==0))

        for det in res["details"]:
            sec = det["section"]
            d, c_ = det["debit"], det["credit"]
            net = None
            if d is not None and c_ is not None: net = round(d+c_, 2)
            elif d is not None: net = d
            elif c_ is not None: net = c_
            row_data = base + [sec, det["field_original"], det["field_en"], get_cn(det["field_en"]),
                               d, c_, net, "", m["source_file"]]
            add_data_row(sec, row_data, SECTION_BG.get(sec,"FFFFFFFF"))

        for chk in res.get("checks", []):
            status = "✓ 核验通过"
            diffs = [chk.get("summary_diff"), chk.get("debit_diff"), chk.get("credit_diff")]
            if any(v is not None and abs(v) > 0.05 for v in diffs):
                status = "核验异常"
            row_data = base + [chk["section"], chk["detail_debits"], chk["detail_credits"],
                               chk["detail_net"], chk["summary_total"], chk["subtotal_debits"],
                               chk["subtotal_credits"], chk["summary_diff"], chk["debit_diff"],
                               chk["credit_diff"], status, m["source_file"]]
            add_check_row(row_data, err=(status != "✓ 核验通过"))

        audit_status = m["filename_audit_status"]
        audit_note = m["filename_audit_note"]
        duplicate_files = duplicate_keys.get((m["store"], m["country_code"], m["year"], m["month"], m["quarter"]), [])
        if len(duplicate_files) > 1:
            audit_status = "疑似重复报告期" if audit_status.startswith("✓") else f"{audit_status}; 疑似重复报告期"
            duplicate_note = "同店铺/站点/年月文件：" + " | ".join(sorted(duplicate_files))
            audit_note = f"{audit_note}; {duplicate_note}" if audit_note else duplicate_note
        audit_row = [m["store"], m["country"], m["country_code"], m["source_file"], m["source_path"], m["period"],
                     m["filename_year"], m["filename_month"], m["period_year"], m["period_month"],
                     m["year"], m["month"], m["quarter"], audit_status, audit_note]
        add_filename_audit_row(audit_row, err=(not audit_status.startswith("✓")))

    widths = [14,12,10,9,44,8,8,8,12,54,48,34,16,16,16,34,28]
    check_widths = [14,12,10,9,44,8,8,8,12,16,16,16,16,16,16,14,14,14,18,28]
    audit_widths = [14,12,10,34,80,44,12,12,12,12,10,10,10,34,80]
    for name in sheet_names:
        ws = sheets[name]
        if name == "Checks":
            these_widths = check_widths
        elif name == "Filename_Audit":
            these_widths = audit_widths
        else:
            these_widths = widths
        for i, w in enumerate(these_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        if ws.max_row >= 2:
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    wb.save(output_path)
    print(f"✅ Excel 已保存：{output_path}")

def build_arg_parser():
    parser = argparse.ArgumentParser(
        description="批量抽取亚马逊交易报表 PDF，并合并为带核验 sheet 的 Excel。",
    )
    parser.add_argument(
        "--base-dir",
        default=BASE_DIR,
        help="PDF 根目录，默认优先使用项目内的“亚马逊汇总报表-PDF（按月）”目录。",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="输出 Excel 路径，默认写入 <base-dir>/outputs/亚马逊交易汇总报表_汇总.xlsx。",
    )
    parser.add_argument(
        "--store",
        action="append",
        default=[],
        help="只处理指定店铺，可重复传入。",
    )
    parser.add_argument(
        "--country",
        action="append",
        default=[],
        help="只处理指定国家/站点，可用中文名或站点代码，可重复传入。",
    )
    parser.add_argument(
        "--quiet",
        action="store_true",
        help="减少处理过程日志。",
    )
    parser.add_argument(
        "--strict",
        action="store_true",
        help="如存在核验警告则以非 0 状态退出。",
    )
    return parser


def resolve_output_path(args):
    if args.output:
        return os.path.abspath(args.output)
    return os.path.join(os.path.abspath(args.base_dir), "outputs", "亚马逊交易汇总报表_汇总.xlsx")


def run(base_dir=BASE_DIR, output_path=None, stores=None, countries=None, quiet=False, strict=False):
    stores = set(stores or [])
    countries = {c.upper() for c in (countries or [])}
    country_names = set(countries)
    for name, code in COUNTRY_NAME_TO_CODE.items():
        if name.upper() in countries or code.upper() in countries:
            country_names.add(name.upper())
            countries.add(code.upper())

    pdfs = collect_pdfs(base_dir)
    if stores:
        pdfs = [p for p in pdfs if p[1] in stores]
    if countries:
        pdfs = [p for p in pdfs if p[2].upper() in country_names or p[3].upper() in countries]

    if not quiet:
        print(f"共发现 {len(pdfs)} 个 PDF，开始处理...\n")
    all_results, warn_files = [], []
    for idx, (pdf_path, store, country, country_code) in enumerate(pdfs, 1):
        fname = os.path.basename(pdf_path)
        if not quiet:
            print(f"[{idx:3d}/{len(pdfs)}] {store}/{country}({country_code})/{fname}", end=" ... ")
        try:
            res = extract_pdf(pdf_path, store, country, country_code)
            all_results.append(res)
            if res["errors"]:
                errs = res["errors"]
                if not quiet:
                    print(f"⚠️  {errs}")
                warn_files.append((fname, res["errors"]))
            else:
                if not quiet:
                    print("✓")
        except Exception as e:
            import traceback; traceback.print_exc()
            if not quiet:
                print(f"❌ {e}")
            warn_files.append((fname, [str(e)]))

    if not quiet:
        print(f"\n成功解析 {len(all_results)} 个文件，写入 Excel...")
    output_path = output_path or os.path.join(os.path.abspath(base_dir), "outputs", "亚马逊交易汇总报表_汇总.xlsx")
    write_excel(all_results, output_path)
    if warn_files:
        if not quiet:
            print(f"\n⚠️  核验警告文件：")
            for f, errs in warn_files:
                print(f"  {f}: {errs}")
    if not quiet:
        print(f"\n完成！输出：{output_path}")
    return 1 if (strict and warn_files) else 0


def main(argv=None):
    parser = build_arg_parser()
    args = parser.parse_args(argv)
    return run(
        base_dir=os.path.abspath(args.base_dir),
        output_path=resolve_output_path(args),
        stores=args.store,
        countries=args.country,
        quiet=args.quiet,
        strict=args.strict,
    )

if __name__ == "__main__":
    sys.exit(main())
