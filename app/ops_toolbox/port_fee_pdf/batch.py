from __future__ import annotations

import re
import traceback
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path

import pandas as pd
import pdfplumber
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


@dataclass
class PortFeePdfJob:
    job_id: str
    source_label: str
    output_dir: Path
    output_path: Path
    summary: dict
    rows: list[dict]
    details: list[dict]


SUMMARY_COLUMNS = [
    "来源文件",
    "Invoice No.",
    "Issue Date",
    "Reprinted Date",
    "Ref No.",
    "Consignee",
    "Vessel",
    "Voyage",
    "ETD Date",
    "Destination",
    "Load Port",
    "Discharge Port",
    "Total Cartons",
    "Total CBM",
    "Master B/L",
    "FCR No.",
    "Remark",
    "Prepared By",
    "Currency",
    "Invoice Amount",
    "费用明细行数",
    "状态",
    "问题说明",
]

DETAIL_BASE_COLUMNS = [
    "来源文件",
    "Invoice No.",
    "Remark",
    "柜号/箱号",
    "Currency",
]

HEADER_EN = {
    "来源文件": "Source File",
    "费用明细行数": "Detail Row Count",
    "状态": "Status",
    "问题说明": "Issues",
    "柜号/箱号": "Container No.",
}

HEADER_CN = {
    "来源文件": "来源文件",
    "Invoice No.": "发票号",
    "Issue Date": "开票日期",
    "Reprinted Date": "重印日期",
    "Ref No.": "参考号",
    "Consignee": "收货人",
    "Vessel": "船名",
    "Voyage": "航次",
    "ETD Date": "预计离港日期",
    "Destination": "目的地",
    "Load Port": "装货港",
    "Discharge Port": "卸货港",
    "Total Cartons": "总箱数",
    "Total CBM": "总体积（立方米）",
    "Master B/L": "主提单号",
    "FCR No.": "货运收据号",
    "Remark": "备注",
    "Prepared By": "制单人",
    "Currency": "币种",
    "Invoice Amount": "发票金额",
    "费用明细行数": "费用明细行数",
    "状态": "状态",
    "问题说明": "问题说明",
    "柜号/箱号": "柜号/箱号",
}

FEE_DESCRIPTION_CN = {
    "CFS RECEIVING CHARGE (CBM)": "CFS 收货费（按立方米）",
    "PORT SECURITY PRO-RATE (CBM)": "港口安保费（按立方米）",
    "SORTING CHARGES": "分拣费",
    "SORTING CHARGE (CBM)": "分拣费（按立方米）",
    "DOCUMENTATION ADMIN FEE (INBOUND)": "进口单证管理费",
    "EXTRA CFS CHARGE (CBM)": "额外 CFS 操作费（按立方米）",
    "ORIGIN CFS LABELING": "起运地 CFS 标签费",
    "SHIFTING CHARGE IN WAREHOUSE": "仓库移位费",
    "STORAGE CHARGE (CBM PER DAY)": "仓储费（按立方米/天）",
    "THC PRO-RATE (CBM)": "码头操作费（按立方米）",
    "VAT OUTPUT PAYABLE/VAT INPUT RECOVERABLE": "应付销项增值税/可抵扣进项增值税",
    "VGM SUBMISSION/FILING FEE (CBM)": "VGM 申报备案费（按立方米）",
    "DOCUMENTATION FEE": "文件费",
    "CUSTOMS DECLARATION FEE": "报关费",
    "CUSTOMS CLEARANCE FEE": "清关费",
    "TERMINAL HANDLING CHARGE": "码头操作费",
    "TERMINAL HANDLING CHARGES": "码头操作费",
    "HANDLING CHARGE": "操作费",
    "HANDLING CHARGES": "操作费",
    "LOADING CHARGE": "装货费",
    "UNLOADING CHARGE": "卸货费",
    "TRUCKING FEE": "拖车费",
    "TRUCKING CHARGE": "拖车费",
    "WAREHOUSE CHARGE": "仓储费",
    "WAREHOUSING CHARGE": "仓储费",
    "INSPECTION FEE": "查验费",
    "MANAGEMENT FEE": "管理费",
}


def process_port_fee_folder(folder: Path, output_dir: Path, job_id: str, label: str | None = None) -> PortFeePdfJob:
    label = label or str(folder)
    output_dir.mkdir(parents=True, exist_ok=True)
    pdf_paths = sorted(path for path in folder.rglob("*") if path.is_file() and path.suffix.lower() == ".pdf")
    rows: list[dict] = []
    details: list[dict] = []

    for pdf_path in pdf_paths:
        try:
            invoice = extract_port_fee_invoice(pdf_path)
            rows.append(invoice["summary"])
            details.extend(invoice["details"])
        except Exception as exc:
            rows.append(
                {
                    "来源文件": pdf_path.name,
                    "Invoice No.": "",
                    "Issue Date": "",
                    "Reprinted Date": "",
                    "Ref No.": "",
                    "Consignee": "",
                    "Vessel": "",
                    "Voyage": "",
                    "ETD Date": "",
                    "Destination": "",
                    "Load Port": "",
                    "Discharge Port": "",
                    "Total Cartons": "",
                    "Total CBM": "",
                    "Master B/L": "",
                    "FCR No.": "",
                    "Remark": "",
                    "Prepared By": "",
                    "Currency": "",
                    "Invoice Amount": "",
                    "费用明细行数": 0,
                    "状态": "解析失败",
                    "问题说明": f"{exc}；{traceback.format_exc(limit=2)}",
                }
            )

    output_path = output_dir / f"港杂费发票汇总_{label_name(label)}.xlsx"
    write_port_fee_excel(rows, details, output_path)
    warning_rows = [row for row in rows if row.get("状态") != "通过"]
    summary = {
        "source_files": len(pdf_paths),
        "processed": len(rows) - sum(1 for row in rows if row.get("状态") == "解析失败"),
        "failed": sum(1 for row in rows if row.get("状态") == "解析失败"),
        "warnings": len(warning_rows),
        "detail_rows": len(details),
        "currency": _single_value(row.get("Currency") for row in rows),
        "total_amount": _sum_amounts(row.get("Invoice Amount") for row in rows),
        "output_filename": output_path.name,
    }
    return PortFeePdfJob(
        job_id=job_id,
        source_label=label,
        output_dir=output_dir,
        output_path=output_path,
        summary=summary,
        rows=rows,
        details=details,
    )


def extract_port_fee_invoice(path: Path) -> dict:
    with pdfplumber.open(path) as pdf:
        page_texts = [page.extract_text() or "" for page in pdf.pages]
        text = "\n".join(page_texts)
        details = []
        for page in pdf.pages:
            details.extend(_extract_detail_lines(page))

    fields = _extract_header_fields(text)
    notes = []
    for key, label in (
        ("invoice_no", "Invoice No."),
        ("issue_date", "Issue Date"),
        ("currency", "Currency"),
        ("invoice_amount", "Invoice Amount"),
    ):
        if not fields.get(key):
            notes.append(f"缺少 {label}")
    if not details:
        notes.append("未识别到费用明细")

    amount_total = _decimal(fields.get("invoice_amount"))
    detail_total = sum((_decimal(item.get("Amount")) or Decimal("0")) for item in details)
    if amount_total is not None and details and abs(detail_total - amount_total) > Decimal("0.02"):
        notes.append(f"费用明细合计 {detail_total} 与发票金额 {amount_total} 不一致")
    summary = {
        "来源文件": path.name,
        "Invoice No.": fields.get("invoice_no", ""),
        "Issue Date": fields.get("issue_date", ""),
        "Reprinted Date": fields.get("reprinted_date", ""),
        "Ref No.": fields.get("ref_no", ""),
        "Consignee": fields.get("consignee", ""),
        "Vessel": fields.get("vessel", ""),
        "Voyage": fields.get("voyage", ""),
        "ETD Date": fields.get("etd_date", ""),
        "Destination": fields.get("destination", ""),
        "Load Port": fields.get("load_port", ""),
        "Discharge Port": fields.get("discharge_port", ""),
        "Total Cartons": fields.get("total_cartons", ""),
        "Total CBM": fields.get("total_cbm", ""),
        "Master B/L": fields.get("master_bl", ""),
        "FCR No.": fields.get("fcr_no", ""),
        "Remark": fields.get("remark", ""),
        "Prepared By": fields.get("prepared_by", ""),
        "Currency": fields.get("currency", ""),
        "Invoice Amount": fields.get("invoice_amount", ""),
        "费用明细行数": len(details),
        "状态": "需复核" if notes else "通过",
        "问题说明": "；".join(notes),
    }
    detail_rows = [
        {
            "来源文件": path.name,
            "Invoice No.": fields.get("invoice_no", ""),
            "Remark": fields.get("remark", ""),
            **item,
            "Currency": fields.get("currency", ""),
        }
        for item in details
    ]
    return {"summary": summary, "details": detail_rows}


def write_port_fee_excel(rows: list[dict], details: list[dict], output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    detail_rows, detail_columns, detail_translations = _pivot_fee_details(details)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pd.DataFrame(rows, columns=SUMMARY_COLUMNS).to_excel(writer, sheet_name="发票汇总", index=False)
        pd.DataFrame(detail_rows, columns=detail_columns).to_excel(writer, sheet_name="费用明细", index=False)
        _add_bilingual_header(writer.sheets["发票汇总"], SUMMARY_COLUMNS, HEADER_CN)
        _add_bilingual_header(writer.sheets["费用明细"], detail_columns, {**HEADER_CN, **detail_translations})
    return output_path


def _add_bilingual_header(worksheet, columns: list[str], translations: dict[str, str]) -> None:
    worksheet.insert_rows(2)
    worksheet.freeze_panes = "A3"
    last_column = get_column_letter(len(columns))
    worksheet.auto_filter.ref = f"A2:{last_column}{worksheet.max_row}"
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for column_index, column in enumerate(columns, start=1):
        english_header = HEADER_EN.get(column, column)
        chinese_header = translations.get(column, column)
        worksheet.cell(row=1, column=column_index, value=english_header)
        worksheet.cell(row=2, column=column_index, value=chinese_header)
        for row_index in (1, 2):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if column in DETAIL_BASE_COLUMNS or column in SUMMARY_COLUMNS:
            width = max(14, min(24, max(len(english_header), len(chinese_header) * 2) + 4))
        else:
            width = 30
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    worksheet.row_dimensions[1].height = 28
    worksheet.row_dimensions[2].height = 28

    for row in worksheet.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)


def _pivot_fee_details(details: list[dict]) -> tuple[list[dict], list[str], dict[str, str]]:
    """将同一张发票的费用项横向展开，便于核对各项港杂费。"""
    rows_by_invoice: dict[tuple[str, str, str, str], dict] = {}
    fee_columns: set[str] = set()
    fee_translations: dict[str, str] = {}

    for detail in details:
        source_file = str(detail.get("来源文件", ""))
        invoice_no = str(detail.get("Invoice No.", ""))
        remark = str(detail.get("Remark", ""))
        currency = str(detail.get("Currency", ""))
        key = (source_file, invoice_no, remark, currency)
        row = rows_by_invoice.setdefault(
            key,
            {
                "来源文件": source_file,
                "Invoice No.": invoice_no,
                "Remark": remark,
                "柜号/箱号": [],
                "Currency": currency,
            },
        )

        containers = str(detail.get("柜号/箱号", "")).strip()
        if containers:
            row["柜号/箱号"].extend(part for part in containers.split("、") if part)

        english_name = _fee_name_without_container(str(detail.get("费用描述", "")))
        chinese_name = _translate_fee_description(english_name)
        column = english_name or "Unrecognized Fee Item"
        fee_columns.add(column)
        fee_translations[column] = chinese_name if english_name else "未识别费用项"
        amount = _decimal(detail.get("Amount")) or Decimal("0")
        row[column] = (row.get(column) or Decimal("0")) + amount

    result = []
    for row in rows_by_invoice.values():
        row["柜号/箱号"] = "、".join(sorted(set(row["柜号/箱号"])))
        for column in fee_columns:
            if column in row:
                row[column] = _format_decimal(row[column])
            else:
                row[column] = ""
        result.append(row)
    return result, [*DETAIL_BASE_COLUMNS, *sorted(fee_columns)], fee_translations


def _fee_name_without_container(description: str) -> str:
    cleaned = re.sub(r"\s*\([A-Z]{4}\d{7}\)", "", description).strip()
    normalized = re.sub(r"\s+", " ", cleaned).upper()
    if normalized.startswith("(") and re.search(r"\)\s*GES$", normalized):
        return "SORTING CHARGES"
    return cleaned


def _translate_fee_description(description: str) -> str:
    normalized = re.sub(r"\s+", " ", description).strip().upper()
    exact = FEE_DESCRIPTION_CN.get(normalized)
    if exact:
        return exact

    patterns = (
        (r"DOCUMENTATION.*ADMIN.*FEE.*INBOUND", "进口单证管理费"),
        (r"DOCUMENTATION.*ADMIN.*FEE", "单证管理费"),
        (r"EXTRA.*CFS.*CHARGE", "额外 CFS 操作费"),
        (r"ORIGIN.*CFS.*LABEL", "起运地 CFS 标签费"),
        (r"PORT.*SECURITY", "港口安保费"),
        (r"SHIFTING.*WAREHOUSE", "仓库移位费"),
        (r"SORTING.*CHARGE", "分拣费"),
        (r"STORAGE.*CHARGE", "仓储费"),
        (r"THC.*PRO.?RATE", "码头操作费"),
        (r"VAT.*OUTPUT.*VAT.*INPUT", "应付销项增值税/可抵扣进项增值税"),
        (r"VGM.*(?:SUBMISSION|FILING)", "VGM 申报备案费"),
        (r"CFS.*RECEIVING", "CFS 收货费"),
        (r"CUSTOMS.*DECLARATION", "报关费"),
        (r"CUSTOMS.*CLEARANCE", "清关费"),
        (r"TRUCK", "拖车费"),
        (r"WAREHOUS", "仓储费"),
        (r"INSPECTION", "查验费"),
        (r"HANDLING", "操作费"),
        (r"UNLOADING", "卸货费"),
        (r"LOADING", "装货费"),
    )
    for pattern, chinese_name in patterns:
        if re.search(pattern, normalized):
            suffix = "（按立方米）" if "CBM" in normalized and "立方米" not in chinese_name else ""
            return f"{chinese_name}{suffix}"
    return "其他港杂费（请核对英文原文）"


def _format_decimal(value: Decimal) -> str:
    return format(value.quantize(Decimal("0.01")), "f")


def label_name(label: str) -> str:
    cleaned = Path(label).name if "/" in label or "\\" in label else label
    cleaned = re.sub(r'[<>:"/\\|?*]', "-", cleaned.strip())
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" .-")
    return cleaned or "批次"


def _extract_header_fields(text: str) -> dict:
    fields = {
        "invoice_no": _match(text, r"INVOICE\s+NO\.\s*:\s*([A-Z0-9-]+)"),
        "issue_date": _match(text, r"ISSUE\s+DATE\s*:\s*(\d{2}/\d{2}/\d{4})"),
        "reprinted_date": _match(text, r"REPRINTED\s+(\d{2}/\d{2}/\d{4})"),
        "ref_no": _match(text, r"REF\s+NO\.\s*:\s*([A-Z0-9-]+)"),
        "consignee": _match(text, r"CONSIGNEE\s*:\s*(.+)"),
        "master_bl": _match(text, r"MASTER\s+B/L\s*:\s*([A-Z0-9-]+)"),
        "fcr_no": _match(text, r"FCR\s+NO\.\s*:\s*([A-Z0-9-]+)"),
        "remark": _match(text, r"Remark\s*:\s*([A-Z0-9-]+)"),
        "prepared_by": _match(text, r"Prepared\s+By\s*:\s*(.+)"),
        "currency": _match(text, r"Invoice\s+Amount\s*:\s*([A-Z]{3})\s*[-\d,.]+"),
        "invoice_amount": _match(text, r"Invoice\s+Amount\s*:\s*[A-Z]{3}\s*([-\d,.]+)"),
    }
    vessel_line = _match(text, r"VESSEL\s*:\s*(.+)")
    if vessel_line:
        vessel_match = re.search(r"(.+?)\s+VOYAGE\s*:\s*(.+)$", vessel_line)
        if vessel_match:
            fields["vessel"] = vessel_match.group(1).strip()
            fields["voyage"] = vessel_match.group(2).strip()
        else:
            fields["vessel"] = vessel_line
            fields["voyage"] = ""
    etd_line = _match(text, r"ETD\s+DATE\s*:\s*(.+)")
    if etd_line:
        etd_match = re.search(r"(.+?)\s+DESTINATION\s*:\s*(.+)$", etd_line)
        if etd_match:
            fields["etd_date"] = etd_match.group(1).strip()
            fields["destination"] = etd_match.group(2).strip()
    port_line = _match(text, r"LOAD\s+PORT\s*:\s*(.+)")
    if port_line:
        port_match = re.search(r"(.+?)\s+DISCHARGE\s+PORT\s*:\s*(.+)$", port_line)
        if port_match:
            fields["load_port"] = port_match.group(1).strip()
            fields["discharge_port"] = port_match.group(2).strip()
    total_match = re.search(r"TOTAL\s*:\s*([\d,]+)\s*\(([-\d,.]+)\s*CBM\)", text)
    if total_match:
        fields["total_cartons"] = total_match.group(1).replace(",", "")
        fields["total_cbm"] = total_match.group(2).replace(",", "")
    return fields


def _extract_detail_lines(page) -> list[dict]:
    words = page.extract_words(x_tolerance=1, y_tolerance=3, keep_blank_chars=False)
    rows = _group_words_by_line(words)
    details = []
    for row in rows:
        top = row[0]["top"]
        if top < 320:
            continue
        row_text = " ".join(word["text"] for word in row)
        if "Invoice Amount" in row_text:
            break
        desc_words = [word["text"] for word in row if word["x0"] < 285]
        qty_words = [word["text"] for word in row if 285 <= word["x0"] < 360]
        unit_words = [word["text"] for word in row if 360 <= word["x0"] < 430]
        price_words = [word["text"] for word in row if 430 <= word["x0"] < 505]
        amount_words = [word["text"] for word in row if word["x0"] >= 505]
        if not desc_words or not amount_words:
            continue
        amount = _last_decimal_text(amount_words)
        if amount is None:
            continue
        description, note = _normalize_description(" ".join(desc_words))
        containers = sorted(set(re.findall(r"[A-Z]{4}\d{7}", description)))
        details.append(
            {
                "费用描述": description,
                "柜号/箱号": "、".join(containers),
                "Quantity": _first_decimal_text(qty_words) or "",
                "Unit": " ".join(unit_words),
                "Unit Price": _first_decimal_text(price_words) or "",
                "Amount": amount,
                "备注": note,
            }
        )
    return details


def _group_words_by_line(words: list[dict]) -> list[list[dict]]:
    grouped: list[list[dict]] = []
    for word in sorted(words, key=lambda item: (item["top"], item["x0"])):
        if not grouped or abs(grouped[-1][0]["top"] - word["top"]) > 3:
            grouped.append([word])
        else:
            grouped[-1].append(word)
    return [sorted(row, key=lambda item: item["x0"]) for row in grouped]


def _normalize_description(description: str) -> tuple[str, str]:
    cleaned = re.sub(r"\s+", " ", description).strip()
    container = _match(cleaned, r"([A-Z]{4}\d{7})")
    suffix = f" ({container})" if container else ""
    if cleaned.startswith("(CHFMS") and "CHARGE (CBM)" in cleaned:
        return f"CFS RECEIVING CHARGE (CBM){suffix}", ""
    if "SHOLR" in cleaned and "RGES" in cleaned:
        return f"SORTING CHARGES{suffix}", ""
    return cleaned, ""


def _match(text: str, pattern: str) -> str:
    matched = re.search(pattern, text, flags=re.IGNORECASE | re.MULTILINE)
    return matched.group(1).strip() if matched else ""


def _first_decimal_text(values: list[str]) -> str | None:
    for value in values:
        if _decimal(value) is not None:
            return value.replace(",", "")
    return None


def _last_decimal_text(values: list[str]) -> str | None:
    for value in reversed(values):
        if _decimal(value) is not None:
            return value.replace(",", "")
    return None


def _decimal(value: object) -> Decimal | None:
    if value is None:
        return None
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _sum_amounts(values) -> str:
    total = Decimal("0")
    for value in values:
        total += _decimal(value) or Decimal("0")
    return f"{total:.2f}"


def _single_value(values) -> str:
    items = sorted({str(value) for value in values if value})
    if not items:
        return ""
    if len(items) == 1:
        return items[0]
    return "多币种"
